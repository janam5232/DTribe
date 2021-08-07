import openpyxl as op
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import zipfile as zp
import glob as gb
import shutil as su
from datetime import datetime as dt
import io
import timeit
import mysql.connector as sql

#DATABASE CONNECTION

dbName = 'AnalystData'
host = 'db1.cs3pcss752yf.ap-south-1.rds.amazonaws.com'
user = 'admin'
password = 'janam12345'

mydb = sql.connect(
    host = "db1.cs3pcss752yf.ap-south-1.rds.amazonaws.com",
    user = "admin",
    passwd = "janam12345",
    database = "AnalystData"
)

cur = mydb.cursor()
print(mydb)

#getting the current directories locaiton (where the data will be)
zipPath = "/home/ec2-user"

#accessing the zip file
file = gb.glob(os.path.join(zipPath, "*.zip"))

#extracting the zip file
zip = zp.ZipFile(file[0])
zip.extractall("/opt/eVolume/temp")

zip.close()

zipPath = "/opt/eVolume/temp/"
#checking if files already exist
#need to change directory paths
outputPath = "/opt/eVolume/dataout/"
curDate = dt.today()
fileName = "model_" + curDate.strftime("%Y%m%d") + ".xlsx"
if os.path.isfile(outputPath + fileName):
    os.remove(outputPath + fileName)
    print("fileRemoved")

curDate = dt.today()
outputFileName = "model_" + curDate.strftime("%Y%m%d")
print(outputFileName)
outputWorkbookPath = outputPath + outputFileName+".xlsx"
outputWorkbook = Workbook()
outputWorkbookWorksheet = outputWorkbook.active
outputWorkbookWorksheet['A1'] = 'Date'
outputWorkbookWorksheet['B1'] = 'Ticker'
outputWorkbookWorksheet['C1'] = 'Type'
outputWorkbookWorksheet['D1'] = 'Quarter'
outputWorkbookWorksheet['E1'] = 'Year'
outputWorkbookWorksheet['F1'] = 'Estimated Total Sold'
outputWorkbookWorksheet['G1'] = 'Estimated Sold Maximum'
outputWorkbookWorksheet['H1'] = 'Estimated Sold Minimum'
outputWorkbookWorksheet['I1'] = 'Forecast w/o SA'
outputWorkbookWorksheet['J1'] = 'Forecase w/o Max'
outputWorkbookWorksheet['K1'] = 'Forecast w/o Min'

data = []

files = os.listdir(zipPath)

files = [file for file in files if ".xlsx" in file]

rowCounterEmpirical = 2
rowCounterRegression = 2

tic = timeit.default_timer()
for file in files:
    ticker = file.split(' ')[0]
    with open(zipPath + file, "rb") as f:
        in_file = io.BytesIO(f.read())
    workbook = op.load_workbook(in_file, read_only=True, data_only=True)
    allWorksheetsInTheFile = workbook.sheetnames
    print("Loading: " + file)
    empericalModelSheets = [sheets for sheets in allWorksheetsInTheFile if "Empirical Model" in sheets]
    data = [sheets for sheets in allWorksheetsInTheFile if sheets == "Data"]
    regressionModelSheets = [sheets for sheets in allWorksheetsInTheFile if "Regression Model" in sheets]

    #sheets with emperical model data
    for sheet in empericalModelSheets:
        workableSheet = workbook[sheet]
        print(sheet)
        for row in workableSheet['D1':'D' + str(workableSheet.max_row)]:
            for cellValue in row:
                tempStr = str(cellValue.value)
                if "Estimated total sold" in tempStr and tempStr[-3] == "Q":
                    estimatedTotalSold = workableSheet[get_column_letter(cellValue.column + 2) + str(cellValue.row)]
                    estimatedMaxSold = workableSheet[get_column_letter(cellValue.column + 2) + str(cellValue.row + 1)]
                    estimatedMinSold = workableSheet[get_column_letter(cellValue.column + 2) + str(cellValue.row + 2)]

                    # print(estimatedTotalSold.value)
                    # print(estimatedMaxSold.value)
                    # print(estimatedMinSold.value)

                    outputWorkbookWorksheet["F"+str(rowCounterEmpirical)] = estimatedTotalSold.value
                    # print("F" + str(rowCounterEmpirical))
                    outputWorkbookWorksheet["G"+str(rowCounterEmpirical)] = estimatedMaxSold.value
                    # print("G" + str(rowCounterEmpirical))
                    outputWorkbookWorksheet["H"+str(rowCounterEmpirical)] = estimatedMinSold.value
                    # print("H" + str(rowCounterEmpirical))
                    if sheet[-5:] != "Model" :
                        sheetNameTmp = sheet.split('-')
                        print(sheetNameTmp[1].strip())
                        outputWorkbookWorksheet["C"+str(rowCounterEmpirical)] = sheetNameTmp[1].strip()
                    else:
                        outputWorkbookWorksheet["C"+str(rowCounterEmpirical)] = "Null"
                    rowCounterEmpirical = rowCounterEmpirical + 1
    
    #sheets with regression model data
    for sheet in regressionModelSheets:
        workableSheet = workbook[sheet]
        print(sheet)
        for row in workableSheet["C1":"R" + str(workableSheet.max_row)]:
            for colObj in row:
                tmpValue = str(colObj.value)
                if tmpValue.strip() == "Max":
                    quarter = workableSheet["D" + str(colObj.row - 1)].value
                    yeartmp = workableSheet["C" + str(colObj.row -1)].value
                    year = "20" + yeartmp[-2:]
                    forecastSA = workableSheet[get_column_letter(colObj.column + 1) + str(colObj.row - 1)].value
                    forecastMin = workableSheet[get_column_letter(colObj.column + 1) + str(colObj.row + 1)].value
                    forecastMax = workableSheet[get_column_letter(colObj.column + 1) + str(colObj.row)].value
                    outputWorkbookWorksheet["A"+str(rowCounterRegression)] = curDate.strftime("%Y-%m-%d")
                    # print("A" + str(rowCounterRegression))
                    outputWorkbookWorksheet["D"+str(rowCounterRegression)] = quarter
                    # print("D" + str(rowCounterRegression))
                    outputWorkbookWorksheet["B"+str(rowCounterRegression)] = ticker
                    # print("B" + str(rowCounterRegression))
                    outputWorkbookWorksheet["E"+str(rowCounterRegression)] = year
                    # print("E" + str(rowCounterRegression))
                    outputWorkbookWorksheet["I"+str(rowCounterRegression)] = forecastSA
                    # print("I" + str(rowCounterRegression))
                    outputWorkbookWorksheet["J"+str(rowCounterRegression)] = forecastMax
                    # print("J" + str(rowCounterRegression))
                    outputWorkbookWorksheet["K"+str(rowCounterRegression)] = forecastMin
                    # print("K" + str(rowCounterRegression))
                    if sheet[-5:] != "Model" :
                        sheetNameTmp = sheet.split('-')
                        # print(sheetNameTmp[1].strip())
                        outputWorkbookWorksheet["C"+str(rowCounterRegression)] = sheetNameTmp[1].strip()
                    else:
                        outputWorkbookWorksheet["C"+str(rowCounterRegression)] = "Null"
                    
                    rowCounterRegression = rowCounterRegression + 1

outputWorkbook.save(outputWorkbookPath)

#deleting files after processing

for file in files:
    os.remove(zipPath + file)

with open(outputPath + fileName, "rb") as f:
        in_file = io.BytesIO(f.read())
finalDataFile = op.load_workbook(in_file, read_only=True)
finalSheet = finalDataFile.active

rowValues = []
for row in finalSheet['A2':get_column_letter(finalSheet.max_column) + str(finalSheet.max_row)]:
    curValue = []
    for cellValue in row:
        curValue.append(cellValue.value)
    rowValues.append(tuple(curValue))

q = "INSERT INTO Data(Date, Ticker, Type, Quarter, Year, EstimatedTotalSold, EstimatedMaxSold, EstimatedMinSold, ForecastwoSA, ForecastwoMax, ForecastwoMin) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
cur.executemany(q, rowValues)
mydb.commit()
mydb.close()
print(rowValues)

archiveDestinationPath = "/opt/eVolume/datout/archive/"
su.move(file[0], archiveDestinationPath)

toc = timeit.default_timer()

print(toc-tic)
